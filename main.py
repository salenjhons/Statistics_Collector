import json
import os
import openpyxl
import requests
import time
import config
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
from tkinter import *
from tkinter.ttk import *
import threading
import random

PRIORITIES = config.priorities
REPORT_LIST = config.report

STATUSES = config.statuses
RESTRICTED_SERVICES = config.res_serv

dates_list = []
executors = []
arr = []
fields = []
names = config.names

output = config.path


def gui(credentials, ses, headers):
    window = Tk()
    window.title("Сборщик статистики")
    window.resizable(False, False)
    names_var = StringVar(value=names)
    Label(window, text="Исполнители :").grid(row=0, column=0, pady=5, padx=5)
    Label(window, text="Период :").grid(row=0, column=5, pady=5)
    lbox = Listbox(window, width=20, height=3, selectmode=MULTIPLE, listvariable=names_var)
    lbox.grid(row=1, column=0, columnspan=3, padx=10, sticky=W)
    status = Label(window, text='')
    status.grid(row=3, column=0, pady=5)
    entry_execs = Entry(window)
    entry_execs.grid(row=2, column=0, padx=10, pady=15, sticky=W)

    Label(window, text="Начало :").grid(row=1, column=4)
    entry_begin = Entry(window)
    entry_begin.grid(row=1, column=5, padx=10)
    Label(window, text="Конец :").grid(row=2, column=4)
    entry_end = Entry(window)
    entry_end.grid(row=2, column=5, padx=10)

    btn1 = Button(window,
                  text="За неделю",
                  command=lambda: get_week())
    btn1.grid(row=3, column=5, pady=5, padx=10, sticky=W + E)
    style = Style(window)
    style.layout('text.Horizontal.TProgressbar',
                 [('Horizontal.Progressbar.trough',
                   {'children': [('Horizontal.Progressbar.pbar',
                                  {'side': 'left', 'sticky': 'ns'})],
                    'sticky': 'nswe'}),
                  ('Horizontal.Progressbar.label', {'sticky': ''})])
    style.configure('text.Horizontal.TProgressbar', text='0/0')

    pb = Progressbar(window, orient=HORIZONTAL, mode='determinate', style='text.Horizontal.TProgressbar', length=300)
    pb.grid(row=4, column=0, columnspan=6, padx=10, pady=5, sticky=W + E)

    Button(window,
           text="Собрать",
           command=lambda: threading.Thread(target=collect_statistics,
                                            args=(credentials, ses, lbox, entry_execs, entry_begin, entry_end,
                                                  pb, style, status, headers)).start()) \
        .grid(row=6, column=4, pady=5, sticky=E)

    Button(window, text="Выход",
           command=lambda: os._exit(0)).grid(row=6, pady=5, padx=5, column=5, sticky=W)
    window.mainloop()


def get_entry(lbox, entry):
    input = entry.get()
    if len(lbox.curselection()) > 0 and len(input) == 0:
        for item in lbox.curselection():
            if lbox.get(item) == 'Дежурный':
                executors.append('Дежурный администратор')
            if lbox.get(item) == 'Казьмина':
                executors.append('Казьмина Анастасия Сергеевна')
            if lbox.get(item) == 'Гусев':
                executors.append('Гусев Дмитрий Александрович')
    elif len(input) > 0 and len(lbox.curselection()) == 0:
        for executor in input.split(','):
            executors.append(executor.strip())


def get_dates(entry1, entry2):
    if len(entry1.get()) != 0 and len(entry2.get()) != 0 and len(dates_list) == 0:
        dates_list.append(entry2.get())
        dates_list.append(entry1.get())
        entry1.delete(0, END)
        entry2.delete(0, END)


def get_editors(ses, url, task_id, headers):
    editors = []
    response = ses.get(url + f'tasklifetime?taskid={task_id}', headers=headers)
    life_time = json.loads(response.content).get('TaskLifetimes')
    for element in life_time:
        editors.append(element['Editor'])
    return editors


def get_headers():
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; rv:91.0) Gecko/20100101 Firefox/91.0",
        "Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0",
        "Mozilla/5.0 (X11; Linux x86_64; rv:95.0) Gecko/20100101 Firefox/95.0",
        "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.3",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 OPR/95.0.4635.25"
    ]
    random_user_agent = random.choice(user_agents)
    headers = {
        'User-Agent': random_user_agent,
        'Content-Type': 'application/json'}
    return headers


def get_element_name(id, element_list):
    for element in element_list:
        if id == element['Id']:
            return element['Name']


def collect_statistics(credentials, ses, lbox, entry_execs, entry_begin, entry_end, pb, style, status, headers):
    if len(arr) != 0:
        arr.clear()
    if len(fields) != 0:
        fields.clear()
    get_dates(entry_begin, entry_end)
    get_entry(lbox, entry_execs)
    if len(dates_list) < 2:
        status['text'] = 'Выбери период'
        return
    if len(executors) == 0:
        status['text'] = 'Выбери исполнителей'
        return
    isExecutor = False
    username = credentials[0].strip()
    password = credentials[1].strip()
    url = credentials[2].strip()
    status['text'] = 'Загрузка сервисов...'

    services, services_ids = get_services_list(ses, username, password, url, headers=headers)
    start_date = datetime.strptime(dates_list.pop(), '%d%m%Y').strftime('%Y-%m-%d')
    end_date = datetime.strptime(dates_list.pop(), '%d%m%Y').strftime('%Y-%m-%d')
    dates_list.clear()
    row = 2
    ws, wb, path = config_excel_file(start_date, end_date)
    status['text'] = 'Загрузка заявок...'
    init_response = ses.get(
        url + f'task?CreatedMoreThan={start_date}&CreatedLessThan={end_date}&pagesize=200&ServiceIds='
        + services_ids, headers=headers)
    paginator = json.loads(init_response.content)['Paginator']
    count = 1
    total = paginator['Count']
    for page_number in range(1, paginator['PageCount'] + 1):
        if page_number != 1:
            response = ses.get(url +
                               f'task?CreatedMoreThan={start_date}&CreatedLessThan={end_date}'
                               f'&page={page_number}&pagesize=200&ServiceIds=' + services_ids, headers=headers)
        else:
            response = init_response
        tasks = json.loads(response.content)['Tasks']
        status['text'] = 'Сбор статистики...'
        for task in tasks:
            style.configure('text.Horizontal.TProgressbar', text=f'{total}/{count}')
            pb['value'] += 100 / total
            #      sys.stdout.write(f"\rВсего / Обработано..........................................{total} / {count}")
            time.sleep(0.4)
            #   #print(json.dumps(task, indent=4, sort_keys=False, ensure_ascii=False))
            editors = get_editors(ses, url, task['Id'], headers)
            for executor in executors:
                for editor in editors:
                    if executor == editor:
                        isExecutor = True
                        break
                if isExecutor == True or executor == task['Creator'] or executor == task['Executors']:
                    service = get_element_name(task['ServiceId'], services)
                    priority = get_element_name(task['PriorityId'], PRIORITIES)
                    category = task['Categories']
                    #   reporting(service, category, arr)
                    reporting(service, category)
                    fields.append(task['Id'])
                    fields.append(task['Created'])
                    fields.append(priority)
                    fields.append(task['Name'])
                    fields.append(service)
                    fields.append(category)
                    fields.append(executor)
                    write_field_common(row, fields, ws, wb, path)
                    fields.clear()
                    row += 1
                isExecutor = False
            count += 1
            pb.update()
    pb['value'] = 0
    style.configure('text.Horizontal.TProgressbar', text='0/0')
    status['text'] = 'Статистика собрана!'
    wb.close()
    create_result_file(start_date, end_date, arr)
    ses.close()
    time.sleep(1)
    os._exit(0)


def config_excel_file(start, end):
    path = output + f'Статистика_{start}-{end}.xlsx'

    try:
        if os.path.exists(path):
            os.remove(path)
        wb = openpyxl.load_workbook(filename=path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb['Sheet']
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 38
    ws.column_dimensions['G'].width = 35
    ws.cell(row=1, column=1, value='№')
    ws.cell(row=1, column=2, value='Создана')
    ws.cell(row=1, column=3, value='Приоритет')
    ws.cell(row=1, column=4, value='Наименование заявки')
    ws.cell(row=1, column=5, value='Сервис')
    ws.cell(row=1, column=6, value='Категория')
    ws.cell(row=1, column=7, value='Исполнитель')

    wb.save(path)
    return ws, wb, path


def return_element(arr):
    for element in range(0, len(arr)):
        return arr[element][2]


def create_result_file(first, second, arr):
    row = 2
    column = 1
    total = 0
    path = output + f'Итог_{first}-{second}.xlsx'

    try:
        if os.path.exists(path):
            os.remove(path)
        wb = openpyxl.load_workbook(filename=path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb['Sheet']
    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.cell(row=1, column=1, value='Сервис')
    ws.cell(row=1, column=2, value='Категория')
    ws.cell(row=1, column=3, value='Количество')

    # for all ids
    for field in REPORT_LIST:
        total += field[2]
        for item in field:
            ws.cell(row=row, column=column, value=item)
            column += 1
        column = 1
        row += 1
    ws.cell(row=row, column=2, value="Всего")
    ws.cell(row=row, column=3, value=total)
    wb.save(filename=path)

    wb.close()


def reporting(service, category):
    if category == '' or category == 'null':
        category = None
    for count, element in enumerate(REPORT_LIST):
        if element[0] == service and element[1] == category:
            element[2] += 1
            return
        else:
            if count + 1 == len(REPORT_LIST):
                REPORT_LIST.append([service, category, 1])
                return


def write_field_common(row, fields, ws, wb, path):
    column = 1
    for field in fields:
        ws.cell(row=row, column=column, value=field)
        column += 1
    wb.save(filename=path)


def get_week():
    yesterday = datetime.today() - timedelta(days=1)
    previous_date = yesterday - timedelta(days=6)
    dates_list.append(yesterday.strftime('%d%m%Y'))
    dates_list.append(previous_date.strftime('%d%m%Y'))


def get_services_list(ses, username, password, url, headers):
    services_ids = ''

    response = ses.get(url + 'service?pagesize=200&fields=Id,Name&for=filtertasks',
                       auth=HTTPBasicAuth(username, password), headers=headers)
    services_list = json.loads(response.content)['Services']
    # json.dumps(services_list, indent=4, sort_keys=False, ensure_ascii=False)
    paginator = json.loads(response.content)['Paginator']
    for page_number in range(1, paginator['PageCount'] + 1):
        response = ses.get(url + f'service?fields=Id,Name&for=filtertasks&page={page_number}&pagesize=200',
                           headers=headers)
        #        print(response.content)
        services_list += (json.loads(response.content)['Services'])
        for element in json.loads(response.content)['Services']:
            service_id = element['Id']
            for count in range(0, len(RESTRICTED_SERVICES)):
                if service_id == RESTRICTED_SERVICES[count]['Id']:
                    break
                else:
                    if count + 1 == len(RESTRICTED_SERVICES):
                        services_ids += str(service_id) + ","
                        break
    return services_list, services_ids[:len(services_ids) - 1]


def get_task(credentials, ses, task_id):
    username = credentials[0].strip()
    password = credentials[1].strip()
    url = credentials[2].strip()
    headers = {
        'Content-Type': 'application/json',
        'User-Agent': 'Mozilla/5.0'}
    response = ses.get(url + f'task?Id={task_id}',
                       auth=HTTPBasicAuth(username, password),
                       headers=headers
                       )

    return json.loads(response.content)['Task']


if __name__ == '__main__':
    credentials = config.creds
    ses = requests.Session()
    headers = get_headers()
    gui(credentials, ses, headers)
